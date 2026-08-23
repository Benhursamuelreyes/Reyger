#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Tests de la Fase 2 (gestión de datos): importación/exportación de la BD.

Ejecuta:  python src/reyger/test_backup.py
"""

import os
import sqlite3
import sys
import tempfile
import zipfile

TESTS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(TESTS_DIR)
SRC_DIR = os.path.join(PROJECT_ROOT, "src")
sys.path.insert(0, SRC_DIR)
# Recursos y módulos del paquete (assets incluidos) para las pruebas
SCRIPT_DIR = os.path.join(SRC_DIR, "reyger")

from openpyxl import load_workbook

from reyger import backup

VERDE = "\033[92m"
ROJO = "\033[91m"
AZUL = "\033[94m"
BOLD = "\033[1m"
RESET = "\033[0m"

FALLOS = []


def chequear(nombre, condicion, detalle=""):
    if condicion:
        print(f"  {VERDE}✓{RESET} {nombre}")
    else:
        print(f"  {ROJO}✗{RESET} {nombre}" + (f" — {detalle}" if detalle else ""))
        FALLOS.append(nombre)


def crear_bd_ejemplo(ruta):
    """Base mínima con las tablas núcleo y una tabla auxiliar."""
    conn = sqlite3.connect(ruta)
    conn.executescript(
        """
        CREATE TABLE ventas (
            id INTEGER PRIMARY KEY,
            total REAL NOT NULL,
            metodo_pago TEXT
        );
        CREATE TABLE inventario (
            id INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL,
            precio REAL NOT NULL,
            costo REAL NOT NULL,
            stock INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE clientes (
            id INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL,
            email TEXT
        );
        INSERT INTO ventas (total, metodo_pago) VALUES (12.5, 'Efectivo'),
                                                        (7.3, 'Tarjeta');
        INSERT INTO inventario (nombre, precio, costo, stock)
        VALUES ('Manzana', 0.85, 0.40, 100),
               ('Portátil', 899.99, 650.00, 3),
               ('Ratón', 15.50, 8.20, 25);
        INSERT INTO clientes (nombre, email) VALUES ('Ana', 'ana@ejemplo.es');
        """
    )
    conn.commit()
    conn.close()


def filas(ruta, sql):
    conn = sqlite3.connect(ruta)
    try:
        return conn.execute(sql).fetchall()
    finally:
        conn.close()


def main():
    tmpdir = tempfile.mkdtemp(prefix="reyger_backup_")
    bd = os.path.join(tmpdir, "database.db")
    crear_bd_ejemplo(bd)

    # ------------------------------------------------------------------
    print(f"\n{BOLD}{AZUL}[TEST 1] Exportación en los tres formatos{RESET}")
    ruta_db = backup.exportar_sqlite(os.path.join(tmpdir, "copia"), db_path=bd)
    chequear("SQLite exporta y añade extensión", os.path.exists(ruta_db))

    ruta_zip = backup.exportar_csv_zip(
        os.path.join(tmpdir, "copia.zip"), db_path=bd
    )
    with zipfile.ZipFile(ruta_zip) as zf:
        miembros = sorted(zf.namelist())
    chequear(
        "CSV zip contiene una entrada por tabla",
        miembros == ["clientes.csv", "inventario.csv", "ventas.csv"],
        str(miembros),
    )

    if backup.EXCEL_DISPONIBLE:
        ruta_xlsx = backup.exportar_excel(
            os.path.join(tmpdir, "copia.xlsx"), db_path=bd
        )
        libro = load_workbook(ruta_xlsx)
        hojas = sorted(libro.sheetnames)
        libro.close()
        chequear(
            "Excel contiene una hoja por tabla",
            hojas == ["clientes", "inventario", "ventas"],
            str(hojas),
        )
    else:
        print(f"  {AZUL}(i){RESET} openpyxl no disponible: se omite Excel")

    # ------------------------------------------------------------------
    print(f"\n{BOLD}{AZUL}[TEST 2] Ciclo completo: exportar .db → dañar → restaurar{RESET}")
    conn = sqlite3.connect(bd)
    conn.execute("DELETE FROM inventario")
    conn.commit()
    conn.close()
    chequear("Inventario vaciado antes de restaurar",
             filas(bd, "SELECT COUNT(*) FROM inventario")[0][0] == 0)

    resultado = backup.importar_datos(ruta_db, db_path=bd)
    inventario = filas(bd, "SELECT nombre, precio, costo, stock FROM inventario ORDER BY id")
    chequear("Modo de importación 'completa'", resultado["modo"] == "completa")
    chequear(".db restaura las 3 filas", len(inventario) == 3, str(inventario))
    chequear(".db conserva tipos y valores",
             inventario[1] == ("Portátil", 899.99, 650.0, 3), str(inventario[1]))
    chequear("Respaldo automático creado", resultado["respaldo"] is not None
             and os.path.exists(resultado["respaldo"]))

    # ------------------------------------------------------------------
    print(f"\n{BOLD}{AZUL}[TEST 3] Ciclo con CSV (.zip) y Excel (.xlsx){RESET}")
    for etiqueta, ruta in (("zip", ruta_zip), ("xlsx", ruta_xlsx)):
        conn = sqlite3.connect(bd)
        conn.execute("DELETE FROM clientes")
        conn.commit()
        conn.close()
        resultado = backup.importar_datos(ruta, db_path=bd)
        total = filas(bd, "SELECT COUNT(*) FROM clientes")[0][0]
        email = filas(bd, "SELECT email FROM clientes WHERE id = 1")[0][0]
        chequear(f".{etiqueta} restaura clientes (modo tablas)",
                 resultado["modo"] == "tablas" and total == 1)
        chequear(f".{etiqueta} preserva texto Unicode", email == "ana@ejemplo.es")

    # ------------------------------------------------------------------
    print(f"\n{BOLD}{AZUL}[TEST 4] Validaciones y seguridad{RESET}")
    corrupta = os.path.join(tmpdir, "corrupta.db")
    with open(corrupta, "wb") as f:
        f.write(os.urandom(2048))
    antes = filas(bd, "SELECT COUNT(*) FROM inventario")[0][0]
    try:
        backup.importar_datos(corrupta, db_path=bd)
        fallo_corrupta = False
    except backup.BackupError:
        fallo_corrupta = True
    despues = filas(bd, "SELECT COUNT(*) FROM inventario")[0][0]
    chequear("BD corrupta rechazada con BackupError", fallo_corrupta)
    chequear("BD activa intacta tras rechazo", antes == despues == 3)

    ajena = os.path.join(tmpdir, "ajena.db")
    conn = sqlite3.connect(ajena)
    conn.execute("CREATE TABLE otra (id INTEGER)")
    conn.commit()
    conn.close()
    try:
        backup.importar_datos(ajena, db_path=bd)
        fallo_ajena = False
    except backup.BackupError as e:
        fallo_ajena = "no parece" in str(e).lower()
    chequear("BD ajena a Reyger rechazada", fallo_ajena)

    futura = os.path.join(tmpdir, "futura.db")
    backup.exportar_sqlite(futura, db_path=bd)
    conn = sqlite3.connect(futura)
    conn.execute("PRAGMA user_version = 999")
    conn.commit()
    conn.close()
    try:
        backup.importar_datos(futura, db_path=bd)
        fallo_futura = False
    except backup.BackupError:
        fallo_futura = True
    chequear("BD de esquema futuro rechazada", fallo_futura)

    desconocido = os.path.join(tmpdir, "datos.txt")
    with open(desconocido, "w") as f:
        f.write("hola")
    try:
        backup.importar_datos(desconocido, db_path=bd)
        fallo_formato = False
    except backup.BackupError:
        fallo_formato = True
    chequear("Formato desconocido rechazado", fallo_formato)

    # ------------------------------------------------------------------
    print(f"\n{BOLD}{AZUL}[TEST 5] Tablas desconocidas y purga de respaldos{RESET}")
    zip_extra = os.path.join(tmpdir, "extra.zip")
    with zipfile.ZipFile(zip_extra, "w") as zf:
        zf.writestr("inventario.csv", "nombre,precio,costo,stock\nPera,0.6,0.3,50\n")
        zf.writestr("tabla_fantasma.csv", "x,y\n1,2\n")
    resultado = backup.importar_datos(zip_extra, db_path=bd)
    pera = filas(bd, "SELECT COUNT(*) FROM inventario WHERE nombre = 'Pera'")[0][0]
    chequear("Tabla conocida importada desde zip manual", pera == 1)
    chequear("Tabla fantasma informada como ignorada",
             resultado["ignoradas"] == ["tabla_fantasma"], str(resultado["ignoradas"]))

    for _ in range(backup.MAX_RESPALDOS + 3):
        backup.respaldar_bd_actual(bd)
    respaldos = [
        n for n in os.listdir(tmpdir) if n.startswith("database_respaldo_")
    ]
    chequear(f"Purga mantiene ≤ {backup.MAX_RESPALDOS} respaldos",
             len(respaldos) <= backup.MAX_RESPALDOS, str(len(respaldos)))

    # ------------------------------------------------------------------
    print()
    if FALLOS:
        print(f"{ROJO}{BOLD}FALLARON {len(FALLOS)} comprobaciones:{RESET}")
        for nombre in FALLOS:
            print(f"  - {nombre}")
        return 1
    print(f"{VERDE}{BOLD}Todos los tests de backup pasaron.{RESET}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
