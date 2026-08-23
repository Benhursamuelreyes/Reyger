"""Importación y exportación de la base de datos de Reyger.

Permite al usuario llevarse sus datos (o restaurarlos) en tres formatos:

- **SQLite** (``.db``/``.sqlite``/``.sqlite3``): copia binaria completa y
  fiel de la base, realizada con el API de backup de SQLite (segura aunque
  existan conexiones abiertas).
- **Excel** (``.xlsx``): un libro con una hoja por tabla. Requiere
  ``openpyxl``.
- **CSV comprimido** (``.zip``): un fichero CSV por tabla dentro de un zip.

En la importación:

- Un ``.db`` sustituye la base completa; se valida antes con
  ``PRAGMA integrity_check`` y se exige que contenga las tablas nucleo.
- Los formatos ``.zip`` (CSV) y ``.xlsx`` sustituyen únicamente el
  *contenido* de las tablas reconocidas; las tablas desconocidas se
  ignoran y se informan.

Antes de cualquier importación se genera automáticamente una copia de
seguridad de la base actual (``database_respaldo_YYYYMMDD_HHMMSS.bak``)
junto a ella; se conservan las :data:`MAX_RESPALDOS` más recientes.

Todas las funciones aceptan ``db_path`` para operar sobre otra base
(útil en tests); por defecto trabajan sobre la del usuario.
"""

import csv
import io
import os
import shutil
import sqlite3
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

try:
    from openpyxl import Workbook, load_workbook

    EXCEL_DISPONIBLE = True
except ImportError:  # pragma: no cover - depende del entorno
    EXCEL_DISPONIBLE = False

from .db import close as cerrar_conexion_compartida
from .resources import get_db_path

MAX_RESPALDOS = 5

EXTENSIONES_SQLITE = (".db", ".sqlite", ".sqlite3")

#: Etiquetas de formato para la interfaz (valor -> descripción/extensión).
FORMATOS_EXPORTACION = {
    "db": ("Base de datos SQLite", ".db"),
    "excel": ("Libro de Excel", ".xlsx"),
    "csv": ("CSV comprimido", ".zip"),
}

TABLAS_NUCLEO = {"ventas", "inventario"}


class BackupError(Exception):
    """Error validado durante una importación/exportación."""


# ---------------------------------------------------------------------------
# Utilidades internas
# ---------------------------------------------------------------------------


def _conectar(db_path=None):
    return sqlite3.connect(str(db_path or get_db_path()))


def _conectar_solo_lectura(ruta):
    """Abre una base en modo lectura pura vía URI (no crea ni recupera)."""
    uri = "file:" + quote(os.path.abspath(str(ruta))) + "?mode=ro"
    try:
        return sqlite3.connect(uri, uri=True)
    except sqlite3.Error as e:
        raise BackupError(f"No se pudo abrir «{ruta}»: {e}") from e


def _tablas(conn):
    filas = conn.execute(
        "SELECT name FROM sqlite_master "
        "WHERE type = 'table' AND name NOT LIKE 'sqlite_%' "
        "ORDER BY name"
    ).fetchall()
    return [fila[0] for fila in filas]


def _leer_tabla(conn, tabla):
    cursor = conn.execute(f'SELECT * FROM "{tabla}"')
    columnas = [d[0] for d in cursor.description]
    filas = cursor.fetchall()
    cursor.close()
    return columnas, filas


def _asegurar_extension(ruta, extension):
    ruta = str(ruta)
    if os.path.splitext(ruta)[1].lower() != extension:
        ruta += extension
    return ruta


def _nombre_hoja(nombre_tabla, usados=None):
    """Nombre válido de hoja Excel: sin []:*?/\\ y de 31 caracteres máximo."""
    hoja = nombre_tabla
    for caracter in '[]:*?/\\':
        hoja = hoja.replace(caracter, "_")
    hoja = hoja[:31] or "Hoja"
    base = hoja
    indice = 2
    while usados is not None and hoja.lower() in usados:
        sufijo = f"~{indice}"
        hoja = base[: 31 - len(sufijo)] + sufijo
        indice += 1
    if usados is not None:
        usados.add(hoja.lower())
    return hoja


# ---------------------------------------------------------------------------
# Exportación
# ---------------------------------------------------------------------------


def exportar_sqlite(destino, db_path=None):
    """Copia binaria completa mediante el API de backup de SQLite."""
    destino = _asegurar_extension(destino, ".db")
    origen_conn = _conectar(db_path)
    try:
        destino_conn = sqlite3.connect(destino)
        try:
            with destino_conn:
                origen_conn.backup(destino_conn)
        finally:
            destino_conn.close()
    finally:
        origen_conn.close()
    return destino


def exportar_csv_zip(destino, db_path=None):
    """Exporta todas las tablas como CSV (una por miembro) dentro de un zip."""
    destino = _asegurar_extension(destino, ".zip")
    conn = _conectar(db_path)
    try:
        tablas = _tablas(conn)
        if not tablas:
            raise BackupError("La base de datos no contiene tablas.")
        with zipfile.ZipFile(destino, "w", zipfile.ZIP_DEFLATED) as zf:
            for tabla in tablas:
                columnas, filas = _leer_tabla(conn, tabla)
                buffer = io.StringIO()
                writer = csv.writer(buffer)
                writer.writerow(columnas)
                for fila in filas:
                    writer.writerow(["" if v is None else v for v in fila])
                zf.writestr(f"{tabla}.csv", buffer.getvalue())
    finally:
        conn.close()
    return destino


def exportar_excel(destino, db_path=None):
    """Exporta todas las tablas a un libro Excel (una hoja por tabla)."""
    if not EXCEL_DISPONIBLE:
        raise BackupError(
            "El soporte de Excel requiere el paquete 'openpyxl', que no está "
            "instalado en este equipo."
        )
    destino = _asegurar_extension(destino, ".xlsx")
    conn = _conectar(db_path)
    try:
        usados = set()
        libro = Workbook(write_only=True)
        try:
            for tabla in _tablas(conn):
                hoja = libro.create_sheet(_nombre_hoja(tabla, usados))
                columnas, filas = _leer_tabla(conn, tabla)
                hoja.append(columnas)
                for fila in filas:
                    hoja.append(list(fila))
            libro.save(destino)
        finally:
            libro.close()
    finally:
        conn.close()
    return destino


def exportar_datos(destino, db_path=None):
    """Exporta eligiendo el formato por la extensión de ``destino``."""
    extension = os.path.splitext(str(destino))[1].lower()
    if extension in EXTENSIONES_SQLITE:
        return exportar_sqlite(destino, db_path), "db"
    if extension == ".xlsx":
        return exportar_excel(destino, db_path), "excel"
    if extension == ".zip":
        return exportar_csv_zip(destino, db_path), "csv"
    raise BackupError(
        f"Extensión no reconocida: «{extension or destino}». "
        "Use .db, .xlsx o .zip."
    )


# ---------------------------------------------------------------------------
# Respaldos automáticos previos a importar
# ---------------------------------------------------------------------------


def respaldar_bd_actual(db_path=None):
    """Copia la base actual a ``database_respaldo_<fecha>.bak``.

    Devuelve la ruta del respaldo o ``None`` si la base aún no existe.
    Conserva solo las :data:`MAX_RESPALDOS` copias más recientes.
    """
    origen = Path(str(db_path or get_db_path()))
    if not origen.exists():
        return None
    sello = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    destino = origen.parent / f"database_respaldo_{sello}.bak"
    shutil.copy2(origen, destino)
    _purgar_respaldos(origen.parent)
    return str(destino)


def _purgar_respaldos(carpeta):
    respaldos = sorted(carpeta.glob("database_respaldo_*.bak"))
    for viejo in respaldos[: max(0, len(respaldos) - MAX_RESPALDOS)]:
        try:
            viejo.unlink()
        except OSError:
            pass


# ---------------------------------------------------------------------------
# Importación
# ---------------------------------------------------------------------------


def _validar_bd_origen(ruta):
    try:
        conn = _conectar_solo_lectura(ruta)
    except sqlite3.Error as e:
        raise BackupError(f"«{ruta}» no es una base de datos válida ({e}).") from e
    try:
        resultado = conn.execute("PRAGMA integrity_check").fetchone()[0]
        if resultado != "ok":
            raise BackupError(
                f"La base de datos está corrupta (integrity_check: {resultado})."
            )
        presentes = set(_tablas(conn))
        faltantes = TABLAS_NUCLEO - presentes
        if faltantes:
            raise BackupError(
                "El fichero no parece una base de datos de Reyger: faltan "
                "las tablas " + ", ".join(sorted(faltantes)) + "."
            )
        from .migrations import LATEST_VERSION

        version = conn.execute("PRAGMA user_version").fetchone()[0]
        if version > LATEST_VERSION:
            raise BackupError(
                "La base pertenece a una versión más moderna de Reyger "
                f"(esquema {version} > {LATEST_VERSION}) y no puede "
                "importarse en esta instalación."
            )
    except sqlite3.DatabaseError as e:
        raise BackupError(
            f"«{ruta}» no es una base de datos SQLite válida ({e})."
        ) from e
    finally:
        conn.close()


def importar_sqlite(ruta_origen, db_path=None):
    """Sustituye la base completa por ``ruta_origen`` tras validarla.

    Cierra la conexión compartida antes de sobrescribir (imprescindible en
    Windows) y aplica las migraciones pendientes en el acto; si estas
    fallan, restaura automáticamente el estado anterior.
    """
    destino = Path(str(db_path or get_db_path()))
    _validar_bd_origen(ruta_origen)
    respaldo = respaldar_bd_actual(str(destino))
    cerrar_conexion_compartida()

    temporal = destino.with_name(destino.name + ".importando")
    shutil.copyfile(str(ruta_origen), temporal)

    from .migrations import run_migrations

    try:
        conn = sqlite3.connect(temporal)
        try:
            run_migrations(conn)
        finally:
            conn.close()
    except Exception as e:
        temporal.unlink(missing_ok=True)
        raise BackupError(
            "La base importada no pudo actualizarse al esquema actual y no "
            f"se aplicó ningún cambio ({e})."
        ) from e

    os.replace(temporal, destino)
    return respaldo


def _leer_csv_zip(ruta):
    datos = {}
    try:
        with zipfile.ZipFile(ruta) as zf:
            for nombre in zf.namelist():
                if not nombre.lower().endswith(".csv"):
                    continue
                tabla = os.path.splitext(os.path.basename(nombre))[0]
                with zf.open(nombre) as f:
                    lector = csv.reader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    filas = list(lector)
                if filas:
                    datos[tabla] = filas
    except zipfile.BadZipFile as e:
        raise BackupError("El fichero .zip no es válido o está dañado.") from e
    if not datos:
        raise BackupError("El zip no contiene ningún fichero CSV.")
    return datos


def _leer_excel(ruta):
    if not EXCEL_DISPONIBLE:
        raise BackupError(
            "El soporte de Excel requiere el paquete 'openpyxl', que no está "
            "instalado en este equipo."
        )
    try:
        libro = load_workbook(str(ruta), read_only=True, data_only=True)
    except Exception as e:
        raise BackupError(f"No se pudo abrir el libro de Excel: {e}") from e
    try:
        datos = {}
        for hoja in libro.worksheets:
            filas = [list(fila) for fila in hoja.iter_rows(values_only=True)]
            while filas and all(v is None for v in filas[-1]):
                filas.pop()
            if filas and any(v is not None for v in filas[0]):
                datos[hoja.title] = filas
    finally:
        libro.close()
    if not datos:
        raise BackupError("El libro de Excel no contiene datos exportables.")
    return datos


def importar_tablas(datos, db_path=None):
    """Sustituye el contenido de las tablas reconocidas en ``datos``.

    ``datos`` mapea ``{tabla: [[cabecera], [fila], ...]}``. Las columnas se
    emparejan por nombre; las que no existan en el esquema actual se
    descartan y las tablas desconocidas se informan. Todo se ejecuta en una
    única transacción: si algo falla, no se altera nada.
    """
    destino = str(db_path or get_db_path())
    respaldo = respaldar_bd_actual(destino)
    conn = sqlite3.connect(destino)
    resumen = {}
    ignoradas = []
    try:
        existentes = set(_tablas(conn))
        for tabla in datos:
            if tabla not in existentes:
                ignoradas.append(tabla)

        conn.execute("PRAGMA foreign_keys = OFF")
        try:
            for tabla, filas in datos.items():
                if tabla not in existentes or len(filas) < 2:
                    continue
                cabecera = [str(c) for c in filas[0]]
                reales = [
                    fila[1]
                    for fila in conn.execute(f'PRAGMA table_info("{tabla}")')
                ]
                mapa = [(i, c) for i, c in enumerate(cabecera) if c in reales]
                if not mapa:
                    ignoradas.append(tabla)
                    continue
                columnas_sql = ", ".join(f'"{c}"' for _, c in mapa)
                marcadores = ", ".join("?" for _ in mapa)
                conn.execute(f'DELETE FROM "{tabla}"')
                insertada = 0
                for fila in filas[1:]:
                    valores = [
                        None if fila[i] == "" else fila[i] for i, _ in mapa
                    ]
                    conn.execute(
                        f'INSERT INTO "{tabla}" ({columnas_sql}) '
                        f"VALUES ({marcadores})",
                        valores,
                    )
                    insertada += 1
                resumen[tabla] = insertada
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.execute("PRAGMA foreign_keys = ON")
    except sqlite3.Error as e:
        raise BackupError(f"Error importando datos: {e}") from e
    finally:
        conn.close()
    return resumen, ignoradas, respaldo


def importar_datos(ruta_origen, db_path=None):
    """Importa ``ruta_origen`` detectando el formato por su extensión.

    Devuelve un diccionario normalizado::

        {"modo": "completa"|"tablas", "respaldo": str|None,
         "resumen": {tabla: filas}, "ignoradas": [tablas]}
    """
    ruta_origen = str(ruta_origen)
    if not os.path.exists(ruta_origen):
        raise BackupError(f"No existe el fichero «{ruta_origen}».")

    extension = os.path.splitext(ruta_origen)[1].lower()
    if extension in EXTENSIONES_SQLITE:
        respaldo = importar_sqlite(ruta_origen, db_path)
        return {
            "modo": "completa",
            "respaldo": respaldo,
            "resumen": {},
            "ignoradas": [],
        }
    if extension == ".zip":
        datos = _leer_csv_zip(ruta_origen)
    elif extension == ".xlsx":
        datos = _leer_excel(ruta_origen)
    else:
        raise BackupError(
            f"Formato no reconocido («{extension}»). Use .db, .sqlite, "
            ".sqlite3, .xlsx o .zip."
        )

    resumen, ignoradas, respaldo = importar_tablas(datos, db_path)
    return {
        "modo": "tablas",
        "respaldo": respaldo,
        "resumen": resumen,
        "ignoradas": ignoradas,
    }
